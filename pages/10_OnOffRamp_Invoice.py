# -*- coding: utf-8 -*-
"""
On/Off-Ramp Invoice Generator + FX Spot Pricer
Tab 1: Bloomberg-style FX margin pricing calculator
Tab 2: Purchase Confirmation XLSX generator
"""

import io
import os
import math
from datetime import date, datetime, timedelta

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------
# Page Config
# --------------------------------------------------
st.set_page_config(page_title="On/Off-Ramp", layout="wide")

# --------------------------------------------------
# Auth
# --------------------------------------------------
APP_PASSWORD = st.secrets.get("APP_PASSWORD", os.environ.get("APP_PASSWORD", ""))

if "auth_ok" not in st.session_state:
    st.session_state.auth_ok = False


def show_login():
    st.title("On/Off-Ramp")
    st.markdown("---")
    st.markdown("### Login Required")
    pw = st.text_input("Password", type="password", key="onofframp_pw")
    if st.button("Login", type="primary"):
        if pw == APP_PASSWORD and APP_PASSWORD:
            st.session_state.auth_ok = True
            st.rerun()
        else:
            st.error("Invalid password.")


if not st.session_state.auth_ok:
    show_login()
    st.stop()


# --------------------------------------------------
# Constants
# --------------------------------------------------
ISSUERS = [
    "Presto Solution Hong Kong Limited",
    "Presto Investment (BVI) Ltd.",
]

COMMON_PAIRS = ["USDT/USD", "USDC/USD", "BTC/USDT", "ETH/USDT", "Custom"]


def parse_pair(pair: str) -> tuple[str, str]:
    """Return (base_ccy, quote_ccy) from 'BASE/QUOTE' string."""
    if "/" in pair:
        parts = pair.split("/", 1)
        return parts[0].strip().upper(), parts[1].strip().upper()
    return pair.upper(), "???"


# --------------------------------------------------
# Excel builder
# --------------------------------------------------
def build_invoice(
    reference: str,
    trade_date: date,
    counterparty_name: str,
    counterparty_addr1: str,
    counterparty_addr2: str,
    attention_name: str,
    attention_email: str,
    direction: str,
    base_amount: float,
    spot_rate: float,
    markup_bps: float,
    contract_date: str,
    value_date: str,
    issuer: str,
) -> Workbook:
    """
    direction='on_ramp'  : We sell USD  / We buy  USDT  (Presto perspective)
    direction='off_ramp' : We sell USDT / We buy  USD
    base_amount          : USDT amount always
    outright_rate        : spot_rate adjusted for markup
    """
    outright_rate = round(spot_rate - markup_bps / 10_000, 8)
    usd_amount = round(base_amount * outright_rate, 2)
    markup_text = (
        "No mark up applied for this trade"
        if markup_bps == 0
        else f"{markup_bps:.1f} bps"
    )

    if direction == "on_ramp":
        sell_amount, sell_ccy = usd_amount, "USD"
        buy_amount, buy_ccy = base_amount, "USDT"
    else:
        sell_amount, sell_ccy = base_amount, "USDT"
        buy_amount, buy_ccy = usd_amount, "USD"

    wb = Workbook()
    ws = wb.active
    ws.title = "Confirmation"

    # Column widths
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16

    def cell(
        coord, value=None, bold=False, size=10, color="000000", align_h="left", fmt=None
    ):
        c = ws[coord]
        if value is not None:
            c.value = value
        c.font = Font(bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align_h, vertical="center")
        if fmt:
            c.number_format = fmt
        return c

    thin = Side(style="thin", color="000000")
    grey = Side(style="thin", color="DDDDDD")
    blue = Side(style="thin", color="1F4E79")

    # ---- Header ----
    cell("A1", f"Reference: [{reference}]", bold=True)
    cell("A3", trade_date.strftime("%d %B %Y"))
    cell("A5", counterparty_name, bold=True)
    cell("A6", counterparty_addr1)
    if counterparty_addr2:
        cell("A7", counterparty_addr2)

    cell("A9", "Attention:  ")
    cell("B9", attention_name or attention_email)
    for col in "ABCD":
        ws[f"{col}10"].border = Border(bottom=Side(style="thin", color="CCCCCC"))

    # ---- Title ----
    ws.merge_cells("A11:D11")
    t = ws["A11"]
    t.value = "CONFIRMATION AND SETTLEMENT INSTRUCTIONS FOR FOREIGN EXCHANGE CONTRACT"
    t.font = Font(bold=True, size=11, color="1F4E79")
    t.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[11].height = 30

    cell("A13", "We confirm our foreign exchange contract as follows:")

    # ---- Trade block ----
    for col in "ABCD":
        ws[f"{col}14"].border = Border(bottom=blue)

    trade_rows = [
        (15, "We sell", sell_amount, sell_ccy, "#,##0.00", True),
        (16, "We buy", buy_amount, buy_ccy, "#,##0.00", True),
        (17, "Contract date", contract_date, None, None, False),
        (18, "Value date", value_date, None, None, False),
        (19, "Spot Reference", spot_rate, " (USDT/USD mid-rate)", "0.0000####", False),
        (20, "Markup", markup_text, None, None, False),
        (21, "Outright Rate", outright_rate, None, "0.0000####", False),
    ]

    for r, label, val, extra, fmt, is_bold in trade_rows:
        ws.row_dimensions[r].height = 16
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(size=10)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = Border(bottom=grey)

        vc = ws.cell(row=r, column=3, value=val)
        vc.font = Font(size=10, bold=is_bold)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border = Border(bottom=grey)
        if fmt:
            vc.number_format = fmt

        if extra:
            ec = ws.cell(row=r, column=4, value=extra)
            ec.font = Font(size=10)
            ec.alignment = Alignment(horizontal="left", vertical="center")
            ec.border = Border(bottom=grey)

    # Top/bottom borders for trade block
    for c in range(1, 5):
        ws.cell(row=15, column=c).border = Border(top=thin, bottom=grey)
        ws.cell(row=21, column=c).border = Border(bottom=thin)

    # ---- Signature ----
    cell("A24", "Yours faithfully,")
    cell("A26", issuer, bold=True)

    return wb


# --------------------------------------------------
# Bloomberg-style FX Pricer CSS
# --------------------------------------------------
PRICER_CSS = """
<style>
.bbg-card {
    background: #0a0f1e;
    border: 1px solid #1e3a5f;
    border-radius: 8px;
    padding: 20px 24px;
    margin-bottom: 12px;
}
.bbg-card-highlight {
    background: #0a1a0a;
    border: 1px solid #2a6a2a;
    border-radius: 8px;
    padding: 20px 24px;
    margin-bottom: 12px;
}
.bbg-ccy-badge {
    display: inline-block;
    font-family: 'Courier New', monospace;
    font-size: 0.72rem;
    font-weight: bold;
    letter-spacing: 0.1em;
    padding: 2px 8px;
    border-radius: 3px;
    margin-right: 6px;
    margin-bottom: 8px;
}
.bbg-base-badge  { background: #1e3a5f; color: #7eb8f7; }
.bbg-quote-badge { background: #2a1a00; color: #ffa040; }
.bbg-pair {
    font-family: 'Courier New', monospace;
    font-size: 1.05rem;
    color: #7eb8f7;
    letter-spacing: 0.05em;
    margin-bottom: 4px;
}
.bbg-label {
    font-family: 'Courier New', monospace;
    font-size: 0.75rem;
    color: #5a7a9a;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    margin-bottom: 6px;
}
.bbg-rate {
    font-family: 'Courier New', monospace;
    font-size: 2.2rem;
    font-weight: bold;
    letter-spacing: 0.04em;
}
.bbg-amount {
    font-family: 'Courier New', monospace;
    font-size: 1.7rem;
    font-weight: bold;
}
.bbg-receives {
    font-family: 'Courier New', monospace;
    font-size: 2rem;
    font-weight: bold;
    color: #60e080;
}
.bbg-pays {
    font-family: 'Courier New', monospace;
    font-size: 1.4rem;
    color: #ff9060;
}
.bbg-bid   { color: #ff6060; }
.bbg-mid   { color: #d0d0d0; }
.bbg-offer { color: #60e080; }
.bbg-pnl {
    font-family: 'Courier New', monospace;
    font-size: 1.4rem;
    color: #ffd700;
    font-weight: bold;
}
.bbg-sub {
    font-family: 'Courier New', monospace;
    font-size: 0.78rem;
    color: #7eb8f7;
}
.bbg-spread {
    font-family: 'Courier New', monospace;
    font-size: 0.9rem;
    color: #aaaaaa;
}
.bbg-divider {
    border: none;
    border-top: 1px solid #1e3a5f;
    margin: 12px 0;
}
</style>
"""


def render_pricer(
    pair: str,
    base_ccy: str,
    quote_ccy: str,
    mid: float,
    bps: float,
    amount: float,
    amount_in: str,  # "base" or "quote"
    direction: str,  # "client_buys_base" | "client_sells_base"
):
    spread = mid * bps / 10_000
    bid = mid - spread
    offer = mid + spread

    if direction == "client_buys_base":
        applied_rate = offer
        rate_label = "OFFER"
        rate_class = "bbg-offer"
        # Client pays quote, receives base
        if amount_in == "base":
            client_pays_amt = amount * applied_rate
            client_pays_ccy = quote_ccy
            client_recvs_amt = amount
            client_recvs_ccy = base_ccy
        else:  # amount in quote
            client_pays_amt = amount
            client_pays_ccy = quote_ccy
            client_recvs_amt = amount / applied_rate
            client_recvs_ccy = base_ccy
    else:  # client_sells_base
        applied_rate = bid
        rate_label = "BID"
        rate_class = "bbg-bid"
        # Client pays base, receives quote
        if amount_in == "base":
            client_pays_amt = amount
            client_pays_ccy = base_ccy
            client_recvs_amt = amount * applied_rate
            client_recvs_ccy = quote_ccy
        else:  # amount in quote
            client_pays_amt = amount / applied_rate
            client_pays_ccy = base_ccy
            client_recvs_amt = amount
            client_recvs_ccy = quote_ccy

    # Margin P&L in quote CCY terms
    notional_quote = (
        client_pays_amt if direction == "client_buys_base" else client_recvs_amt
    )
    margin_pnl = notional_quote * bps / 10_000

    st.markdown(PRICER_CSS, unsafe_allow_html=True)

    # ---- Row 1: BID / MID / OFFER ----
    c1, c2, c3 = st.columns(3)
    active_bid = (
        "border: 1px solid #ff6060;" if direction == "client_sells_base" else ""
    )
    active_offer = (
        "border: 1px solid #60e080;" if direction == "client_buys_base" else ""
    )

    with c1:
        st.markdown(
            f"""
<div class="bbg-card" style="{active_bid}">
<div class="bbg-label">BID — client sells {base_ccy}</div>
<div class="bbg-rate bbg-bid">{bid:.6f}</div>
<div class="bbg-spread">−{spread:.6f} from mid</div>
</div>""",
            unsafe_allow_html=True,
        )

    with c2:
        st.markdown(
            f"""
<div class="bbg-card">
<div class="bbg-label">MID &nbsp;·&nbsp; <span style="color:#7eb8f7">{pair}</span></div>
<div class="bbg-rate bbg-mid">{mid:.6f}</div>
<div class="bbg-spread">spread ± {bps:.1f} bps</div>
</div>""",
            unsafe_allow_html=True,
        )

    with c3:
        st.markdown(
            f"""
<div class="bbg-card" style="{active_offer}">
<div class="bbg-label">OFFER — client buys {base_ccy}</div>
<div class="bbg-rate bbg-offer">{offer:.6f}</div>
<div class="bbg-spread">+{spread:.6f} from mid</div>
</div>""",
            unsafe_allow_html=True,
        )

    st.markdown("")

    # ---- Row 2: Trade Summary (big) + P&L ----
    cs, cp = st.columns([3, 2])

    with cs:
        direction_txt = (
            f"Client BUYS {base_ccy} (pays {quote_ccy})"
            if direction == "client_buys_base"
            else f"Client SELLS {base_ccy} (receives {quote_ccy})"
        )
        pays_fmt = f"{client_pays_amt:,.6f}".rstrip("0").rstrip(".")
        recvs_fmt = f"{client_recvs_amt:,.6f}".rstrip("0").rstrip(".")

        st.markdown(
            f"""
<div class="bbg-card-highlight">
<span class="bbg-ccy-badge bbg-base-badge">BASE: {base_ccy}</span>
<span class="bbg-ccy-badge bbg-quote-badge">QUOTE: {quote_ccy}</span>
<div class="bbg-label" style="margin-top:6px">{direction_txt} &nbsp;·&nbsp; applied: {rate_label} @ <span class="{rate_class}">{applied_rate:.6f}</span></div>
<hr class="bbg-divider">
<div class="bbg-label">CLIENT PAYS</div>
<div class="bbg-pays">{pays_fmt} <span style="color:#ffa040">{client_pays_ccy}</span></div>
<hr class="bbg-divider">
<div class="bbg-label">CLIENT RECEIVES</div>
<div class="bbg-receives">{recvs_fmt} <span style="color:#7ef7a0">{client_recvs_ccy}</span></div>
</div>""",
            unsafe_allow_html=True,
        )

    with cp:
        st.markdown(
            f"""
<div class="bbg-card">
<div class="bbg-label">Margin P&L</div>
<div class="bbg-pnl">{quote_ccy} {margin_pnl:,.4f}</div>
<div class="bbg-sub">{notional_quote:,.2f} {quote_ccy} × {bps:.1f} bps</div>
<hr class="bbg-divider">
<div class="bbg-label">Applied Rate</div>
<div class="bbg-amount {rate_class}">{applied_rate:.6f}</div>
<div class="bbg-sub">{rate_label} &nbsp;|&nbsp; spread {spread:.8f}</div>
</div>""",
            unsafe_allow_html=True,
        )

    return {
        "applied_rate": applied_rate,
        "rate_label": rate_label,
        "client_pays_amt": client_pays_amt,
        "client_pays_ccy": client_pays_ccy,
        "client_recvs_amt": client_recvs_amt,
        "client_recvs_ccy": client_recvs_ccy,
        "margin_pnl": margin_pnl,
        "spread": spread,
    }


# --------------------------------------------------
# Main UI
# --------------------------------------------------
st.title("On/Off-Ramp")
st.markdown("FX Spot Pricer + Purchase Confirmation Generator")
st.markdown("---")

tab1, tab2 = st.tabs(["FX Spot Pricer", "Invoice Generator"])


# ========================================
# TAB 1: FX SPOT PRICER
# ========================================
with tab1:
    st.markdown("### FX Spot Margin Pricer")
    st.caption("Bloomberg-style bid/offer calculator with margin P&L")
    st.markdown("")

    p_col1, p_col2 = st.columns([1, 2], gap="large")

    with p_col1:
        pair_choice = st.selectbox(
            "Currency Pair", COMMON_PAIRS, index=0, key="pricer_pair"
        )
        if pair_choice == "Custom":
            pair_label = st.text_input(
                "Custom pair (BASE/QUOTE)", value="XXX/YYY", key="custom_pair"
            )
        else:
            pair_label = pair_choice

        base_ccy, quote_ccy = parse_pair(pair_label)

        # CCY badges
        st.markdown(
            f'<span style="display:inline-block;background:#1e3a5f;color:#7eb8f7;'
            f"font-family:monospace;font-size:0.8rem;font-weight:bold;padding:3px 10px;"
            f'border-radius:4px;margin-right:6px">BASE: {base_ccy}</span>'
            f'<span style="display:inline-block;background:#2a1a00;color:#ffa040;'
            f"font-family:monospace;font-size:0.8rem;font-weight:bold;padding:3px 10px;"
            f'border-radius:4px">QUOTE: {quote_ccy}</span>',
            unsafe_allow_html=True,
        )
        st.markdown("")

        mid_price = st.number_input(
            "Mid Price",
            min_value=0.000001,
            value=0.9988,
            step=0.0001,
            format="%.6f",
            key="pricer_mid",
        )

        margin_bps = st.number_input(
            "Margin (bps)",
            min_value=0.0,
            max_value=1000.0,
            value=5.0,
            step=0.5,
            format="%.1f",
            key="pricer_bps",
        )

        st.markdown("**Trade Amount**")
        amt_ccy_choice = st.radio(
            "Amount denomination",
            ["base", "quote"],
            format_func=lambda x: (
                f"In {base_ccy} (base)" if x == "base" else f"In {quote_ccy} (quote)"
            ),
            horizontal=True,
            key="pricer_amt_ccy",
        )
        trade_amount = st.number_input(
            f"Amount ({base_ccy if amt_ccy_choice == 'base' else quote_ccy})",
            min_value=0.0,
            value=100_000.0,
            step=1_000.0,
            format="%.2f",
            key="pricer_amount",
        )

        direction_choice = st.radio(
            "Client Direction",
            ["client_buys_base", "client_sells_base"],
            format_func=lambda x: (
                f"Client BUYS {base_ccy} → OFFER applied"
                if x == "client_buys_base"
                else f"Client SELLS {base_ccy} → BID applied"
            ),
            key="pricer_direction",
        )

    with p_col2:
        result = render_pricer(
            pair=pair_label,
            base_ccy=base_ccy,
            quote_ccy=quote_ccy,
            mid=mid_price,
            bps=margin_bps,
            amount=trade_amount,
            amount_in=amt_ccy_choice,
            direction=direction_choice,
        )
        applied_rate = result["applied_rate"]
        client_pays_amt = result["client_pays_amt"]
        client_pays_ccy = result["client_pays_ccy"]
        client_recvs_amt = result["client_recvs_amt"]
        client_recvs_ccy = result["client_recvs_ccy"]

        st.markdown("")
        if st.button("Use this rate in Invoice Generator →", key="push_to_invoice"):
            st.session_state["invoice_spot_rate"] = applied_rate
            st.session_state["invoice_bps"] = 0.0
            st.success(
                f"Rate {applied_rate:.6f} pushed to Invoice tab. Switch to Invoice Generator."
            )

        # ---- Telegram Message ----
        st.markdown("---")
        st.markdown("#### Telegram Message")

        rate_str = f"{applied_rate:.6f}".rstrip("0").rstrip(".")
        pays_str = f"{client_pays_amt:,.2f} {client_pays_ccy}"
        recvs_str = f"{client_recvs_amt:,.2f} {client_recvs_ccy}"

        tg_msg = (
            f"Rate: {rate_str}\n" f"You send: {pays_str}\n" f"You receive: {recvs_str}"
        )

        st.code(tg_msg, language=None)
        st.caption("우측 상단 복사 아이콘으로 복붙")


# ========================================
# TAB 2: INVOICE GENERATOR
# ========================================
with tab2:
    st.markdown("### Purchase Confirmation Generator")
    st.caption("Generates FX confirmation XLSX (Purchase Confirmation template)")
    st.markdown("")

    i_left, i_right = st.columns([1, 1], gap="large")

    with i_left:
        st.markdown("#### Trade Details")

        inv_direction = st.radio(
            "Direction",
            ["on_ramp", "off_ramp"],
            format_func=lambda x: (
                "On-Ramp  (Presto sells USD / buys USDT)"
                if x == "on_ramp"
                else "Off-Ramp (Presto sells USDT / buys USD)"
            ),
            horizontal=False,
            key="inv_direction",
        )

        inv_usdt = st.number_input(
            "USDT Amount",
            min_value=0.0,
            value=1_000.0,
            step=100.0,
            format="%.2f",
            key="inv_usdt",
        )

        # Pre-fill from pricer if available
        default_rate = st.session_state.get("invoice_spot_rate", 0.9988)
        inv_spot = st.number_input(
            "Spot Reference (USDT/USD mid-rate)",
            min_value=0.9,
            max_value=1.01,
            value=float(default_rate),
            step=0.0001,
            format="%.6f",
            key="inv_spot",
        )

        inv_no_markup = st.checkbox("No markup", value=True, key="inv_no_markup")
        if inv_no_markup:
            inv_bps = 0.0
        else:
            inv_bps = st.number_input(
                "Markup (bps)",
                min_value=0.0,
                max_value=500.0,
                value=float(st.session_state.get("invoice_bps", 5.0)),
                step=1.0,
                format="%.1f",
                key="inv_bps_input",
            )

        outright = inv_spot - inv_bps / 10_000
        usd_equiv = round(inv_usdt * outright, 2)
        if inv_direction == "on_ramp":
            st.info(
                f"We sell **{usd_equiv:,.2f} USD** / We buy **{inv_usdt:,.2f} USDT**  |  Rate: {outright:.6f}"
            )
        else:
            st.info(
                f"We sell **{inv_usdt:,.2f} USDT** / We buy **{usd_equiv:,.2f} USD**  |  Rate: {outright:.6f}"
            )

        st.markdown("#### Dates")
        inv_trade_date = st.date_input("Trade Date", value=date.today(), key="inv_td")

        if st.checkbox("Contract date = Trade date", value=True, key="inv_cd_same"):
            contract_str = inv_trade_date.strftime("%d %B %Y")
        else:
            contract_str = st.text_input(
                "Contract Date", placeholder="28 April 2026", key="inv_cd_manual"
            )

        if st.checkbox("Value date = T+1", value=True, key="inv_vd_same"):
            value_str = (inv_trade_date + timedelta(days=1)).strftime("%d %B %Y")
        else:
            value_str = st.text_input(
                "Value Date", placeholder="29 April 2026", key="inv_vd_manual"
            )

    with i_right:
        st.markdown("#### Reference")
        inv_year = datetime.today().strftime("%Y")
        inv_num = st.number_input(
            "Invoice #", min_value=1, max_value=9999, value=1, step=1, key="inv_num"
        )
        inv_ref = st.text_input(
            "Reference (editable)", value=f"TC-{inv_year}-{inv_num:03d}", key="inv_ref"
        )

        st.markdown("#### Counterparty")
        cp_name = st.text_input(
            "Name", value="Presto Investment (BVI) Ltd.", key="cp_name"
        )
        cp_addr1 = st.text_input(
            "Address Line 1", value="Intershore Chambers, Road Town", key="cp_addr1"
        )
        cp_addr2 = st.text_input(
            "Address Line 2", value="Tortola, British Virgin Islands", key="cp_addr2"
        )
        attn_name = st.text_input("Attention (name)", value="", key="cp_attn")
        attn_mail = st.text_input("Attention (email)", value="", key="cp_email")

        st.markdown("#### Issuer")
        issuer = st.selectbox("Issuing Entity", ISSUERS, index=0, key="inv_issuer")

    st.markdown("---")
    if st.button(
        "Generate Invoice", type="primary", use_container_width=True, key="inv_generate"
    ):
        if inv_usdt <= 0:
            st.error("USDT Amount must be > 0.")
        elif not cp_name.strip():
            st.error("Counterparty name is required.")
        else:
            wb = build_invoice(
                reference=inv_ref,
                trade_date=inv_trade_date,
                counterparty_name=cp_name,
                counterparty_addr1=cp_addr1,
                counterparty_addr2=cp_addr2,
                attention_name=attn_name,
                attention_email=attn_mail,
                direction=inv_direction,
                base_amount=inv_usdt,
                spot_rate=inv_spot,
                markup_bps=inv_bps,
                contract_date=contract_str,
                value_date=value_str,
                issuer=issuer,
            )

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            fname = f"PurchaseConfirmation_{inv_ref}_{inv_trade_date.strftime('%Y%m%d')}.xlsx"

            st.success(f"Generated: **{fname}**")
            st.download_button(
                "Download XLSX",
                data=buf,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="inv_download",
            )

            outright_final = inv_spot - inv_bps / 10_000
            usd_final = round(inv_usdt * outright_final, 2)
            sell_str = (
                f"{usd_final:,.2f} USD"
                if inv_direction == "on_ramp"
                else f"{inv_usdt:,.2f} USDT"
            )
            buy_str = (
                f"{inv_usdt:,.2f} USDT"
                if inv_direction == "on_ramp"
                else f"{usd_final:,.2f} USD"
            )

            with st.expander("Preview", expanded=True):
                st.markdown(
                    f"""
| Field | Value |
|-------|-------|
| Reference | `{inv_ref}` |
| Date | {inv_trade_date.strftime('%d %B %Y')} |
| Counterparty | {cp_name} |
| We Sell | {sell_str} |
| We Buy | {buy_str} |
| Spot Reference | {inv_spot:.6f} USDT/USD |
| Markup | {"No markup" if inv_bps == 0 else f"{inv_bps:.1f} bps"} |
| Outright Rate | {outright_final:.6f} |
| Contract Date | {contract_str} |
| Value Date | {value_str} |
| Issued by | {issuer} |
"""
                )
