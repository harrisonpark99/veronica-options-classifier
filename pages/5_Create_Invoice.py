# -*- coding: utf-8 -*-
"""
Create Invoice - Auto-generate invoice from trade execution CSV
Generates Excel invoice with same cell positions as template
"""

import io
import os
import math
from datetime import datetime, timezone, timedelta
from typing import Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# Page Config
# -----------------------------
st.set_page_config(page_title="Create Invoice", layout="wide")


# -----------------------------
# Auth
# -----------------------------
APP_PASSWORD = st.secrets.get("APP_PASSWORD", os.environ.get("APP_PASSWORD", ""))

if "auth_ok" not in st.session_state:
    st.session_state.auth_ok = False


def show_login():
    """Display login form for this page."""
    st.title("Create Invoice")
    st.markdown("---")
    st.markdown("### Login Required")

    pw = st.text_input(
        "Password",
        type="password",
        placeholder="Enter password",
        key="create_invoice_password"
    )

    if st.button("Login", type="primary"):
        if pw == APP_PASSWORD and APP_PASSWORD:
            st.session_state.auth_ok = True
            st.rerun()
        else:
            st.error("Invalid password or password not configured.")


if not st.session_state.auth_ok:
    show_login()
    st.stop()


# -----------------------------
# Binance CSV Configuration
# -----------------------------
BINANCE_REQUIRED_COLS = [
    "symbol", "id", "orderId", "orderListId", "price", "qty", "quoteQty",
    "commission", "commissionAsset", "time", "isBuyer", "isMaker", "isBestMatch",
]
BINANCE_NUMERIC_COLS = ["price", "qty", "quoteQty", "commission"]

TABLE_HEADERS = [
    "symbol", "id", "orderId", "orderListId", "price", "qty", "quoteQty",
    "commission", "commissionAsset", "time", "isBuyer", "isMaker", "isBestMatch",
    "fill_time(UTC+8)", "fill_value", "sum_fill_amount", "sum_fill_value", "ave_fill_price"
]  # A~R (18 cols)


# -----------------------------
# Helper Functions
# -----------------------------
def compute_totals(df: pd.DataFrame) -> Tuple[float, float, float, float]:
    """Compute filled_amount, filled_value, avg_price, total_commission."""
    filled_amount = float(pd.to_numeric(df["qty"], errors="coerce").fillna(0).sum())
    filled_value = float(pd.to_numeric(df["quoteQty"], errors="coerce").fillna(0).sum())
    avg_price = (filled_value / filled_amount) if filled_amount else 0.0
    total_commission = float(pd.to_numeric(df["commission"], errors="coerce").fillna(0).sum())
    return filled_amount, filled_value, avg_price, total_commission


def ensure_fill_time_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Add fill_time(UTC+8), fill_value columns for template compatibility."""
    out = df.copy()

    if "fill_time(UTC+8)" not in out.columns:
        if "time" in out.columns:
            utc = pd.to_datetime(out["time"], unit="ms", utc=True, errors="coerce")
            utc8 = utc.dt.tz_convert("Asia/Shanghai")
            out["fill_time(UTC+8)"] = utc8.dt.strftime("%Y-%m-%d %H:%M:%S")
        else:
            out["fill_time(UTC+8)"] = ""

    if "fill_value" not in out.columns:
        out["fill_value"] = out["quoteQty"]

    for c in ["sum_fill_amount", "sum_fill_value", "ave_fill_price"]:
        if c not in out.columns:
            out[c] = None

    return out


def infer_sheet_date(df: pd.DataFrame, side: str) -> str:
    """Generate sheet name from date: BUY=YYMMDD, SELL=YYYYMMDD."""
    dt = None
    if "fill_time(UTC+8)" in df.columns and df["fill_time(UTC+8)"].notna().any():
        s = str(df["fill_time(UTC+8)"].dropna().iloc[0])
        try:
            dt = datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            dt = None
    if dt is None and "time" in df.columns:
        try:
            ms = int(df["time"].dropna().iloc[0])
            dt = datetime.fromtimestamp(ms / 1000.0, tz=timezone.utc)
            dt = dt.astimezone(timezone(timedelta(hours=8))).replace(tzinfo=None)
        except Exception:
            dt = datetime.now()

    if dt is None:
        dt = datetime.now()

    if side.upper() == "BUY":
        return dt.strftime("%y%m%d")
    return dt.strftime("%Y%m%d")


def set_col_widths(ws, max_col=18):
    """Set column widths for template appearance."""
    widths = {
        1: 12, 2: 16, 3: 16, 4: 14, 5: 12, 6: 12, 7: 14, 8: 12, 9: 14,
        10: 16, 11: 10, 12: 10, 13: 12, 14: 20, 15: 14, 16: 16, 17: 16, 18: 16
    }
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(c, 14)


def style_header(ws, row=14, start_col=1, end_col=18):
    """Apply header styling to table header row."""
    header_font = Font(bold=True, size=10, color="FFFFFF")
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal="center", vertical="center")

    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        cell.alignment = align


def border_table(ws, r1, r2, c1=1, c2=18):
    """Apply borders to table area."""
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


def border_cells(ws, cells: list):
    """Apply borders to specific cells."""
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in cells:
        ws[cell].border = border


def set_white_borders(ws, max_row=100, max_col=100):
    """Set white borders on all cells to make sheet look like clean white paper."""
    white = Side(style="thin", color="FFFFFF")
    white_border = Border(left=white, right=white, top=white, bottom=white)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = white_border


def truncate_usdt(value: float) -> float:
    """Truncate USDT value to 2 decimal places (cut off, not round)."""
    return math.floor(value * 100) / 100


def autofit_columns(ws, min_width: float = 8, max_width: float = 45):
    """
    Auto-fit column widths based on string length of cell values.
    Keeps widths within [min_width, max_width].
    """
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, float):
                s = f"{v:.8f}".rstrip("0").rstrip(".")
            else:
                s = str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


def build_invoice_workbook(
    df_raw: pd.DataFrame,
    client: str,
    side: str,
    fee_rate: float,
    rebate_usdt: float = 0.0,
) -> Workbook:
    """Create invoice Excel workbook matching template cell positions."""
    side = side.upper()
    df = ensure_fill_time_columns(df_raw)

    # Compute totals
    filled_amount, filled_value, avg_price, _ = compute_totals(df)

    # Fee calculations
    if side == "BUY":
        net = filled_value - float(rebate_usdt)
        gross = net / (1.0 - fee_rate) if (1.0 - fee_rate) else net
        fee_amount = gross * fee_rate
    else:  # SELL
        gross = filled_value
        fee_amount = gross * fee_rate
        net = gross - fee_amount

    sheet_date = infer_sheet_date(df, side)
    symbol = df["symbol"].iloc[0] if "symbol" in df.columns and len(df) > 0 else "N/A"
    base_asset = symbol.replace("USDT", "") if symbol != "N/A" else "BASE"

    wb = Workbook()

    # =========================================================
    # Summary Sheet
    # =========================================================
    ws_sum = wb.active
    ws_sum.title = "Summary"

    # Clean white paper look (100 rows x 100 cols)
    set_white_borders(ws_sum, max_row=100, max_col=100)

    ws_sum["A1"] = "Client:"
    ws_sum["B1"] = client
    ws_sum["A1"].font = Font(bold=True)

    ws_sum["A3"] = "Summary"
    ws_sum["A3"].font = Font(bold=True)

    # Layout: Execution summary (A~C), gap at D, Order summary (E~G)
    ws_sum["A5"] = "Execution summary"
    ws_sum["B5"] = "Amount"
    ws_sum["E5"] = "Order summary"
    ws_sum["A5"].font = Font(bold=True)
    ws_sum["E5"].font = Font(bold=True)

    if side == "BUY":
        ws_sum["A6"] = "Filled Amount:"
        ws_sum["B6"] = filled_amount
        ws_sum["C6"] = f"({base_asset})"

        ws_sum["E6"] = "Order type"
        ws_sum["F6"] = "Buy"

        ws_sum["A7"] = "Filled Value:"
        ws_sum["B7"] = filled_value
        ws_sum["C7"] = "(USDT)"

        ws_sum["E7"] = "Buy order amount"
        ws_sum["F7"] = truncate_usdt(gross)
        ws_sum["G7"] = "(USDT)"

        ws_sum["A8"] = "Average Filled Price:"
        ws_sum["B8"] = avg_price
        ws_sum["C8"] = f"(USDT/{base_asset})"

        ws_sum["E8"] = f"Fee({fee_rate*100:.2f}%)"
        ws_sum["F8"] = truncate_usdt(fee_amount)
        ws_sum["G8"] = "(USDT)"

        ws_sum["A9"] = f"Fee ({fee_rate*100:.2f}%)"
        ws_sum["B9"] = truncate_usdt(fee_amount)
        ws_sum["C9"] = "(USDT)"

        ws_sum["E9"] = "Net of fee Buy order amount"
        ws_sum["F9"] = truncate_usdt(gross - fee_amount)
        ws_sum["G9"] = "(USDT)"

        ws_sum["A10"] = "Settlement Amount:"
        ws_sum["B10"] = filled_amount
        ws_sum["C10"] = f"({base_asset})"

        if rebate_usdt and rebate_usdt != 0:
            ws_sum["A11"] = f"*CEX rebate USDT {rebate_usdt:.2f} included"

        border_cells(ws_sum, [
            "A6", "B6", "E6", "F6",
            "A7", "B7", "E7", "F7",
            "A8", "B8", "E8", "F8",
            "A9", "B9", "E9", "F9",
            "A10", "B10",
        ])

    else:  # SELL
        ws_sum["A6"] = "Filled Amount:"
        ws_sum["B6"] = filled_amount
        ws_sum["C6"] = f"({base_asset})"

        ws_sum["E6"] = "Order type"
        ws_sum["F6"] = "Sell"

        ws_sum["A7"] = "Filled Value:"
        ws_sum["B7"] = filled_value
        ws_sum["C7"] = "(USDT)"

        ws_sum["E7"] = "Sell order amount"
        ws_sum["F7"] = truncate_usdt(gross)
        ws_sum["G7"] = "(USDT)"

        ws_sum["A8"] = "Average Filled Price:"
        ws_sum["B8"] = avg_price
        ws_sum["C8"] = f"(USDT/{base_asset})"

        ws_sum["E8"] = f"Fee({fee_rate*100:.2f}%)"
        ws_sum["F8"] = truncate_usdt(fee_amount)
        ws_sum["G8"] = "(USDT)"

        ws_sum["A9"] = f"Fee ({fee_rate*100:.2f}%)"
        ws_sum["B9"] = truncate_usdt(fee_amount)
        ws_sum["C9"] = "(USDT)"

        ws_sum["E9"] = "Net of fee Sell order amount"
        ws_sum["F9"] = truncate_usdt(net)
        ws_sum["G9"] = "(USDT)"

        ws_sum["A10"] = "Settlement Amount:"
        ws_sum["B10"] = truncate_usdt(net)
        ws_sum["C10"] = "(USDT)"

        border_cells(ws_sum, [
            "A6", "B6", "E6", "F6",
            "A7", "B7", "E7", "F7",
            "A8", "B8", "E8", "F8",
            "A9", "B9", "E9", "F9",
            "A10", "B10",
        ])

    # Auto-fit Summary sheet columns
    autofit_columns(ws_sum, min_width=10, max_width=45)

    # =========================================================
    # Date Sheet (Trading Data)
    # =========================================================
    ws = wb.create_sheet(sheet_date)

    ws["A1"] = "Client:"
    ws["B1"] = client
    ws["A1"].font = Font(bold=True)

    ws["A3"] = "Summary"
    ws["A3"].font = Font(bold=True)

    ws["A5"] = "Execution summary"
    ws["B5"] = "Amount"
    ws["E5"] = "Order summary"
    ws["A5"].font = Font(bold=True)
    ws["E5"].font = Font(bold=True)

    if side == "BUY":
        ws["A6"] = "Filled Amount:"
        ws["B6"] = "=P15"
        ws["C6"] = f"({base_asset})"

        ws["E6"] = "Order type"
        ws["F6"] = "Buy"

        ws["A7"] = "Filled Value:"
        ws["B7"] = "=TRUNC(Q15,2)"
        ws["C7"] = "(USDT)"

        ws["E7"] = "Buy order amount"
        ws["F7"] = truncate_usdt(gross)
        ws["G7"] = "(USDT)"

        ws["A8"] = "Average Filled Price:"
        ws["B8"] = "=TRUNC(R15,2)"
        ws["C8"] = f"(USDT/{base_asset})"

        ws["E8"] = f"Fee({fee_rate*100:.2f}%)"
        ws["F8"] = f"=TRUNC(F7*{fee_rate},2)"
        ws["G8"] = "(USDT)"

        ws["A9"] = f"Fee ({fee_rate*100:.2f}%)"
        ws["B9"] = "=F8"
        ws["C9"] = "(USDT)"

        ws["E9"] = "Net of fee Buy order amount"
        ws["F9"] = "=TRUNC(F7-F8,2)"
        ws["G9"] = "(USDT)"

        ws["A10"] = "Settlement Amount:"
        ws["B10"] = "=B6"
        ws["C10"] = f"({base_asset})"

        if rebate_usdt and rebate_usdt != 0:
            ws["A11"] = f"*CEX rebate USDT {rebate_usdt:.2f} included"

        border_cells(ws, [
            "A6", "B6", "E6", "F6",
            "A7", "B7", "E7", "F7",
            "A8", "B8", "E8", "F8",
            "A9", "B9", "E9", "F9",
            "A10", "B10",
        ])

    else:  # SELL
        ws["A6"] = "Filled Amount:"
        ws["B6"] = "=P15"
        ws["C6"] = f"({base_asset})"

        ws["E6"] = "Order type"
        ws["F6"] = "Sell"

        ws["A7"] = "Filled Value:"
        ws["B7"] = "=TRUNC(Q15,2)"
        ws["C7"] = "(USDT)"

        ws["E7"] = "Sell order amount"
        ws["F7"] = truncate_usdt(gross)
        ws["G7"] = "(USDT)"

        ws["A8"] = "Average Filled Price:"
        ws["B8"] = "=TRUNC(R15,2)"
        ws["C8"] = f"(USDT/{base_asset})"

        ws["E8"] = f"Fee({fee_rate*100:.2f}%)"
        ws["F8"] = f"=TRUNC(F7*{fee_rate},2)"
        ws["G8"] = "(USDT)"

        ws["A9"] = f"Fee ({fee_rate*100:.2f}%)"
        ws["B9"] = "=F8"
        ws["C9"] = "(USDT)"

        ws["E9"] = "Net of fee Sell order amount"
        ws["F9"] = "=TRUNC(F7-F8,2)"
        ws["G9"] = "(USDT)"

        ws["A10"] = "Settlement Amount:"
        ws["B10"] = "=F9"
        ws["C10"] = "(USDT)"

        border_cells(ws, [
            "A6", "B6", "E6", "F6",
            "A7", "B7", "E7", "F7",
            "A8", "B8", "E8", "F8",
            "A9", "B9", "E9", "F9",
            "A10", "B10",
        ])

    # Auto-fit for summary area
    autofit_columns(ws, min_width=10, max_width=45)

    # =========================================================
    # Trading Data section
    # =========================================================
    ws["A13"] = "Trading Data"
    ws["A13"].font = Font(bold=True)

    # Header row 14
    for col_idx, h in enumerate(TABLE_HEADERS, start=1):
        ws.cell(row=14, column=col_idx, value=h)
    style_header(ws, row=14, start_col=1, end_col=len(TABLE_HEADERS))

    # Set table column widths
    set_col_widths(ws, max_col=len(TABLE_HEADERS))

    # Prepare data
    df_out = df.copy()
    for h in TABLE_HEADERS:
        if h not in df_out.columns:
            df_out[h] = None
    df_out["fill_value"] = df_out["quoteQty"]

    # Write data rows starting at row 15
    start_row = 15
    n = len(df_out)

    for i in range(n):
        r = start_row + i
        row = df_out.iloc[i]
        for col_idx, h in enumerate(TABLE_HEADERS, start=1):
            val = row[h]
            if pd.isna(val):
                val = None
            ws.cell(row=r, column=col_idx, value=val)

    # Totals in P15/Q15/R15
    ws["P15"] = filled_amount
    ws["Q15"] = filled_value
    ws["R15"] = avg_price

    # Table borders
    border_table(ws, r1=14, r2=start_row + max(n, 1) - 1, c1=1, c2=len(TABLE_HEADERS))

    return wb


# -----------------------------
# Streamlit UI
# -----------------------------
st.title("Create Invoice")
st.caption("Generate invoice from trade execution CSV")

# Sidebar
with st.sidebar:
    st.header("Settings")

    # 1. Order Side
    st.subheader("1. Order Type")
    order_side = st.radio(
        "Select BUY/SELL",
        options=["BUY", "SELL"],
        horizontal=True
    )
    selected_side = order_side

    st.markdown("---")

    # 2. CEX Selection
    st.subheader("2. Exchange")
    cex_option = st.radio(
        "Select CEX",
        options=["Binance", "Others"],
        horizontal=True
    )

    st.markdown("---")

    # 3. Additional Settings (Binance only)
    if cex_option == "Binance":
        st.subheader("3. Details")
        client = st.text_input("Client", value="")
        fee_rate = st.number_input(
            "Fee rate (e.g. 0.25% = 0.0025)",
            value=0.0025,
            step=0.0001,
            format="%.6f"
        )
        if selected_side == "BUY":
            rebate_usdt = st.number_input(
                "CEX Rebate (USDT)",
                value=0.0,
                step=0.01,
                format="%.2f",
                help="CEX rebate amount for BUY orders"
            )
        else:
            rebate_usdt = 0.0

# Main content
if cex_option == "Others":
    st.warning("Currently only Binance is supported. Other exchanges coming soon.")
    st.stop()

# Binance flow
st.subheader("Upload Trade History")
csv_file = st.file_uploader("Trade history CSV file", type=["csv"])

if csv_file:
    try:
        df = pd.read_csv(csv_file)
    except Exception as e:
        st.error(f"Failed to read CSV: {e}")
        st.stop()

    df.columns = [c.strip() for c in df.columns]
    missing = [c for c in BINANCE_REQUIRED_COLS if c not in df.columns]
    if missing:
        st.error(f"Missing required columns (Binance format): {missing}")
        st.stop()

    for col in BINANCE_NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df_prepared = ensure_fill_time_columns(df)
    filled_amount, filled_value, avg_price, total_commission = compute_totals(df_prepared)

    st.subheader("Trade Summary")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Rows", f"{len(df):,}")
    col2.metric("Side", selected_side)
    col3.metric("Filled Amount", f"{filled_amount:,.8f}")
    col4.metric("Filled Value", f"{filled_value:,.2f} USDT")

    st.caption(f"Average Price: {avg_price:,.2f} USDT | Total Commission: {total_commission:,.8f}")

    with st.expander("Preview Trade History", expanded=False):
        st.dataframe(df.head(30), use_container_width=True)

    st.markdown("---")

    if st.button("Generate Invoice", type="primary", use_container_width=True):
        wb = build_invoice_workbook(
            df_raw=df,
            client=client,
            side=selected_side,
            fee_rate=float(fee_rate),
            rebate_usdt=float(rebate_usdt) if selected_side == "BUY" else 0.0,
        )

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        symbol = df["symbol"].iloc[0] if df["symbol"].nunique() == 1 else "multi"
        date_str = datetime.now().strftime("%Y%m%d")
        file_name = f"invoice_{symbol}_{selected_side.lower()}_{date_str}.xlsx"

        st.success("Invoice generated successfully!")
        st.download_button(
            label="Download Invoice XLSX",
            data=out.getvalue(),
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

else:
    st.info("Upload a Binance trade history CSV file to generate an invoice.")
